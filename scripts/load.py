"""Load YAML data → SQLite. Rebuilds data.db from scratch each run.

Every numeric fact row MUST have a source_id and that source_id MUST exist in sources.yaml.
The SQLite FK on source_id enforces this at load time (so a typo = failure, not silent).
"""
from __future__ import annotations

import datetime as dt
import re
import sqlite3
import sys
from pathlib import Path

import yaml

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
DB = ROOT / "data.db"
SCHEMA = ROOT / "schema.sql"


def read_yaml(p: Path):
    with p.open() as f:
        return yaml.safe_load(f)


def timeline_bucket(statement_date: object, date_precision: str) -> str:
    if isinstance(statement_date, (dt.date, dt.datetime)):
        statement_date = statement_date.isoformat()[:10]
    else:
        statement_date = str(statement_date)
    if date_precision == "day":
        m = re.fullmatch(r"(\d{4})-(\d{2})-\d{2}", statement_date)
        if not m:
            raise ValueError(f"day precision requires YYYY-MM-DD, got {statement_date!r}")
        month = int(m.group(2))
        if not 1 <= month <= 12:
            raise ValueError(f"invalid month in statement_date {statement_date!r}")
        return f"{m.group(1)}Q{((month - 1) // 3) + 1}"
    if date_precision == "month":
        m = re.fullmatch(r"(\d{4})-(\d{2})", statement_date)
        if not m:
            raise ValueError(f"month precision requires YYYY-MM, got {statement_date!r}")
        month = int(m.group(2))
        if not 1 <= month <= 12:
            raise ValueError(f"invalid month in statement_date {statement_date!r}")
        return f"{m.group(1)}Q{((month - 1) // 3) + 1}"
    if date_precision == "quarter":
        if not re.fullmatch(r"\d{4}Q[1-4]", statement_date):
            raise ValueError(f"quarter precision requires YYYYQn, got {statement_date!r}")
        return statement_date
    if date_precision == "year":
        if not re.fullmatch(r"\d{4}", statement_date):
            raise ValueError(f"year precision requires YYYY, got {statement_date!r}")
        return statement_date
    if date_precision == "inferred":
        for precision in ("day", "month", "quarter", "year"):
            try:
                return timeline_bucket(statement_date, precision)
            except ValueError:
                continue
        raise ValueError(f"inferred precision still needs a parseable date, got {statement_date!r}")
    raise ValueError(f"unknown date_precision {date_precision!r}")


def rebuild_schema(conn: sqlite3.Connection) -> None:
    conn.executescript("PRAGMA foreign_keys = OFF;")
    # drop all tables + views
    cur = conn.execute(
        "SELECT type, name FROM sqlite_master "
        "WHERE type IN ('table','view') AND name NOT LIKE 'sqlite_%'"
    )
    for t, n in cur.fetchall():
        conn.execute(f"DROP {t} IF EXISTS {n}")
    conn.commit()
    conn.executescript(SCHEMA.read_text())
    conn.execute("PRAGMA foreign_keys = ON;")


def load_sources(conn: sqlite3.Connection) -> set[str]:
    rows = read_yaml(DATA / "sources.yaml")
    for r in rows:
        conn.execute(
            """INSERT INTO sources
               (id, publisher, title, url, pub_date, retrieved_on, kind, supports, notes)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (r["id"], r["publisher"], r["title"], r["url"],
             r.get("pub_date"), r.get("retrieved_on"),
             r.get("kind"), r.get("supports"), r.get("notes")),
        )
    conn.commit()
    return {r["id"] for r in rows}


def load_contracts(conn: sqlite3.Connection) -> int:
    """Load all contract YAMLs. The doc-level `company` field is used as the
    default operator name for hyperscaler files; non-hyperscaler files
    (neoclouds.yaml, colocation.yaml, sovereign.yaml) declare `operator_type`
    at the doc level and per-row `operator` instead.
    """
    n = 0
    for p in sorted((DATA / "contracts").glob("*.yaml")):
        doc = read_yaml(p)
        # Two doc shapes:
        #  (a) hyperscaler files: doc["company"] = 'Microsoft' (one operator), per-row no 'operator'
        #  (b) operator-typed files: doc["operator_type"] = 'AI-Cloud' / 'Colocation' / 'Sovereign',
        #      each row has its own "operator" (the company name) since the file may mix Crusoe / Lambda / etc.
        default_company = doc.get("company")
        default_op_type = doc.get("operator_type", "Hyperscaler")
        for r in doc["rows"]:
            company    = r.get("operator") or default_company
            op_type    = r.get("operator_type", default_op_type)
            conn.execute(
                """INSERT INTO hyperscaler_contracts
                   (company, operator_type, announced_date, year, cod_year, cod_note,
                    generation_type, capacity_mw, storage_power_mw, storage_energy_mwh,
                    confidence, deal_name,
                    counterparty, contract_years, geography, status,
                    connection_type, connection_reason, notes, source_id)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (company, op_type, r.get("announced_date"), r["year"],
                 r.get("cod_year"), r.get("cod_note"),
                 r["generation_type"], r["capacity_mw"],
                 r.get("storage_power_mw"), r.get("storage_energy_mwh"),
                 r.get("confidence", "Estimated"), r["deal_name"],
                 r.get("counterparty"), r.get("contract_years"),
                 r.get("geography", "US"), r.get("status", "Announced"),
                 r.get("connection_type", "Unknown"),
                 r.get("connection_reason"),
                 r.get("notes"), r["source_id"]),
            )
            n += 1
    conn.commit()
    return n


def _safe_float(v):
    """Tolerant float parser — EIA cells sometimes contain ' ' or 'N/A' instead of empty."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.upper() in ('N/A', 'NA', '#N/A'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_planned_generators(conn: sqlite3.Connection) -> int:
    """Load EIA-860M Planned-sheet snapshots as a federal supply-side time series.

    Walks data/external/eia860m_history/*.xlsx — each file is one monthly snapshot.
    Vintage is parsed from filename (e.g. 'march_generator2026.xlsx' → '2026-03').

    Status code maps to a coarse tier. The tier→probability mapping is no longer
    used as the headline number — instead we track how MW migrate between tiers
    over successive snapshots, which is the actual empirical signal.
    """
    try:
        import openpyxl
    except ImportError:
        print("[skip] openpyxl not installed; skipping planned_generators load")
        return 0

    history_dir = ROOT / "data" / "external" / "eia860m_history"
    if not history_dir.exists():
        print(f"[skip] {history_dir} not found")
        return 0

    files = sorted(history_dir.glob("*_generator*.xlsx"))
    if not files:
        print(f"[skip] no eia860m vintages found in {history_dir}")
        return 0

    EIA_TIERS = {
        '(TS)': ('ConstructionComplete', 0.95),
        '(V)':  ('MajorityComplete',     0.85),
        '(U)':  ('MinorityComplete',     0.70),
        '(T)':  ('ApprovalsReceived',    0.50),
        '(L)':  ('ApprovalsPending',     0.30),
        '(P)':  ('PlannedOnly',          0.15),
        '(OT)': ('Other',                0.20),
    }
    MONTHS = {m: f"{i+1:02d}" for i, m in enumerate(
        ['january','february','march','april','may','june',
         'july','august','september','october','november','december'])}
    SOURCE_ID = 'S291'

    n_total = 0
    for f in files:
        # Filename pattern: <month>_generator<YYYY>.xlsx → vintage = YYYY-MM
        stem = f.stem.lower()
        try:
            month_word, year_part = stem.split('_generator')
            vintage = f"{year_part}-{MONTHS[month_word]}"
        except (KeyError, ValueError):
            print(f"[skip] cannot parse vintage from {f.name}")
            continue

        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        if 'Planned' not in wb.sheetnames:
            print(f"[skip] {f.name} has no 'Planned' sheet")
            continue
        ws = wb['Planned']
        rows = ws.iter_rows(values_only=True)

        # Find header row (looks for 'Plant ID' to anchor)
        hdr = None
        for r in rows:
            if r and any(c == 'Plant ID' for c in r):
                hdr = list(r)
                break
        if hdr is None:
            print(f"[skip] no header found in {f.name}")
            continue

        # Map column name → index
        idx = {h: i for i, h in enumerate(hdr) if h}
        col_status = idx.get('Status')
        col_state  = idx.get('Plant State')
        col_county = idx.get('County')
        col_ba     = idx.get('Balancing Authority Code')
        col_sector = idx.get('Sector')
        col_pid    = idx.get('Plant ID')
        col_gid    = idx.get('Generator ID')
        col_pname  = idx.get('Plant Name')
        col_eid    = idx.get('Entity ID')
        col_ename  = idx.get('Entity Name')
        col_net    = idx.get('Net Summer Capacity (MW)')
        col_name   = idx.get('Nameplate Capacity (MW)')
        col_tech   = idx.get('Technology')
        col_fuel   = idx.get('Energy Source Code')
        col_pm     = idx.get('Prime Mover Code')
        col_year   = idx.get('Planned Operation Year')
        col_month  = idx.get('Planned Operation Month')
        col_lat    = idx.get('Latitude')
        col_lon    = idx.get('Longitude')

        n_file = 0
        for r in rows:
            if not r or r[col_pid] is None or r[col_pname] is None or r[col_status] is None:
                continue
            status = str(r[col_status])
            prefix = status[:status.index(')')+1] if ')' in status else '(OT)'
            tier, prob = EIA_TIERS.get(prefix, ('Other', 0.20))
            try:
                conn.execute(
                    """INSERT INTO planned_generators
                       (vintage, planned_year, planned_month, entity_id, entity_name,
                        producer_type, plant_name, plant_state, county, balancing_authority,
                        lat, lon, plant_id, generator_id,
                        net_summer_capacity_mw, nameplate_capacity_mw, technology,
                        energy_source_code, prime_mover_code, status, status_tier,
                        delivery_probability, source_id)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (vintage,
                     int(r[col_year]) if r[col_year] is not None else None,
                     int(r[col_month]) if r[col_month] is not None else None,
                     int(r[col_eid]) if r[col_eid] is not None else None,
                     str(r[col_ename] or '').strip(),
                     str(r[col_sector] or '').strip() if col_sector is not None else None,
                     str(r[col_pname]).strip(),
                     r[col_state],
                     r[col_county] if col_county is not None else None,
                     r[col_ba] if col_ba is not None else None,
                     _safe_float(r[col_lat]) if col_lat is not None else None,
                     _safe_float(r[col_lon]) if col_lon is not None else None,
                     int(r[col_pid]),
                     str(r[col_gid]) if r[col_gid] else None,
                     _safe_float(r[col_net]),
                     _safe_float(r[col_name]),
                     r[col_tech], r[col_fuel], r[col_pm],
                     status, tier, prob, SOURCE_ID),
                )
                n_file += 1
            except sqlite3.IntegrityError:
                # Same generator with different status in same vintage — keep first
                pass
        wb.close()
        n_total += n_file
        print(f"  loaded {n_file:>5d} rows from {vintage} ({f.name})")

    conn.commit()
    return n_total


def load_operating_generators(conn: sqlite3.Connection) -> int:
    """Load EIA-860M Operating sheet (every plant in commercial operation).

    Operating dates are time-invariant — a plant operating in 2018 shows the same
    operating-year in every vintage. So we just take the latest vintage file
    and load its Operating sheet once. This becomes the ground truth for the
    "operated this quarter" metric used in the transitions chart.
    """
    try:
        import openpyxl
    except ImportError:
        return 0

    history_dir = ROOT / "data" / "external" / "eia860m_history"
    if not history_dir.exists():
        return 0
    files = sorted(history_dir.glob("*_generator*.xlsx"))
    if not files:
        return 0

    # Pick the most recent file by parsing its filename's vintage
    MONTHS = {m: i+1 for i, m in enumerate(
        ['january','february','march','april','may','june',
         'july','august','september','october','november','december'])}
    def vintage_of(f):
        try:
            mo, yr = f.stem.lower().split('_generator')
            return int(yr) * 100 + MONTHS[mo]
        except Exception:
            return 0
    latest = max(files, key=vintage_of)

    wb = openpyxl.load_workbook(latest, data_only=True, read_only=True)
    if 'Operating' not in wb.sheetnames:
        return 0
    ws = wb['Operating']

    # Find header row
    rows = ws.iter_rows(values_only=True)
    hdr = None
    for r in rows:
        if r and any(c == 'Plant ID' for c in r):
            hdr = list(r); break
    if hdr is None:
        return 0
    idx = {h: i for i, h in enumerate(hdr) if h}

    n = 0
    for r in rows:
        if not r or r[idx['Plant ID']] is None or r[idx.get('Generator ID', 10)] is None:
            continue
        try:
            conn.execute(
                """INSERT OR IGNORE INTO operating_generators
                   (plant_id, generator_id, entity_name, plant_name, plant_state,
                    county, balancing_authority, sector, technology, energy_source_code,
                    net_summer_capacity_mw, nameplate_capacity_mw,
                    operating_year, operating_month, status, source_id)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (int(r[idx['Plant ID']]),
                 str(r[idx['Generator ID']]),
                 str(r[idx.get('Entity Name', 1)] or '').strip(),
                 str(r[idx['Plant Name']] or '').strip(),
                 r[idx.get('Plant State')],
                 r[idx.get('County')],
                 r[idx.get('Balancing Authority Code')],
                 r[idx.get('Sector')],
                 r[idx.get('Technology')],
                 r[idx.get('Energy Source Code')],
                 _safe_float(r[idx.get('Net Summer Capacity (MW)')]),
                 _safe_float(r[idx.get('Nameplate Capacity (MW)')]),
                 int(r[idx['Operating Year']]) if r[idx.get('Operating Year')] not in (None, '', ' ') else None,
                 int(r[idx['Operating Month']]) if r[idx.get('Operating Month')] not in (None, '', ' ') else None,
                 r[idx.get('Status')],
                 'S291'),
            )
            n += 1
        except (ValueError, TypeError):
            pass
    wb.close()
    conn.commit()
    return n


def load_campuses(conn: sqlite3.Connection) -> int:
    p = DATA / "campuses.yaml"
    if not p.exists():
        return 0
    doc = read_yaml(p)
    n = 0
    for r in doc["rows"]:
        conn.execute(
            """INSERT INTO data_center_campuses
               (campus_id, campus_name, hyperscaler, primary_tenant,
                city, state_or_region, country, lat, lon,
                capacity_definition, it_load_mw_planned, it_load_mw_phase1,
                it_load_mw_energized, cod_phase1_year, cod_full_year,
                status, power_source_summary, primary_use, notes, source_id)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["campus_id"], r["campus_name"], r["hyperscaler"],
             r.get("primary_tenant"), r.get("city"), r.get("state_or_region"),
             r.get("country", "US"), r.get("lat"), r.get("lon"),
             r.get("capacity_definition", "Critical-IT"),
             r.get("it_load_mw_planned"), r.get("it_load_mw_phase1"),
             r.get("it_load_mw_energized", 0),
             r.get("cod_phase1_year"), r.get("cod_full_year"),
             r.get("status", "Announced"),
             r.get("power_source_summary"), r.get("primary_use"),
             r.get("notes"), r["source_id"]),
        )
        n += 1
    conn.commit()
    return n


def load_campus_evidence(conn: sqlite3.Connection) -> int:
    p = DATA / "campus_evidence.yaml"
    if not p.exists():
        return 0
    doc = read_yaml(p) or {}
    n = 0
    for r in doc.get("rows", []):
        conn.execute(
            """INSERT INTO campus_evidence
               (evidence_id, campus_id, evidence_date, evidence_type,
                independence_group, claim_status, capacity_definition,
                capacity_mw, phase, collectability, evidence_strength,
                quote, notes, source_id)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["evidence_id"], r["campus_id"], r["evidence_date"],
             r["evidence_type"], r["independence_group"],
             r["claim_status"],
             r.get("capacity_definition", "Unknown"),
             r.get("capacity_mw"), r.get("phase"),
             r.get("collectability", "Manual"),
             r["evidence_strength"], r.get("quote"),
             r.get("notes"), r["source_id"]),
        )
        n += 1
    conn.commit()
    return n


def load_proxy_signal_definitions(conn: sqlite3.Connection) -> int:
    p = DATA / "proxy_signal_definitions.yaml"
    if not p.exists():
        return 0
    doc = read_yaml(p) or {}
    n = 0
    for r in doc.get("rows", []):
        metric_keys = r.get("metric_keys", [])
        if isinstance(metric_keys, list):
            metric_keys = ",".join(metric_keys)
        conn.execute(
            """INSERT INTO proxy_signal_definitions
               (signal_id, signal_name, signal_group, metric_keys, source_route,
                update_frequency, confidence, validates, cannot_validate, notes)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (r["signal_id"], r["signal_name"], r["signal_group"],
             metric_keys, r["source_route"], r["update_frequency"],
             r["confidence"], r["validates"], r["cannot_validate"],
             r.get("notes")),
        )
        n += 1
    conn.commit()
    return n


def load_sec_proxy_metrics(conn: sqlite3.Connection) -> int:
    p = DATA / "sec_proxy_metrics.yaml"
    if not p.exists():
        return 0
    doc = read_yaml(p) or {}
    n = 0
    for r in doc.get("rows", []):
        conn.execute(
            """INSERT INTO sec_proxy_metrics
               (ticker, cik, company_name, company_group, metric_key, metric_label,
                taxonomy, xbrl_tag, unit, period_type, period_year, sec_fiscal_year, fiscal_period,
                form, filed_date, start_date, end_date, frame, accession, value,
                source_url, source_id, retrieved_on)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["ticker"], r["cik"], r["company_name"], r["company_group"],
             r["metric_key"], r["metric_label"], r["taxonomy"], r["xbrl_tag"],
             r["unit"], r["period_type"], r["period_year"], r.get("sec_fiscal_year"),
             r.get("fiscal_period") or "", r.get("form") or "",
             r.get("filed_date"), r.get("start_date"), r["end_date"],
             r.get("frame"), r.get("accession") or "", r["value"],
             r["source_url"], r["source_id"], r["retrieved_on"]),
        )
        n += 1
    conn.commit()
    return n


def load_sec_filing_text_signals(conn: sqlite3.Connection) -> int:
    p = DATA / "sec_filing_text_signals.yaml"
    if not p.exists():
        return 0
    doc = read_yaml(p) or {}
    n = 0
    for r in doc.get("rows", []):
        conn.execute(
            """INSERT INTO sec_filing_text_signals
               (ticker, cik, company_name, company_group, form, filed_date,
                report_date, accession, document_url, signal_type, matched_term,
                snippet, source_id, retrieved_on)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["ticker"], r["cik"], r["company_name"], r["company_group"],
             r["form"], r["filed_date"], r.get("report_date"),
             r["accession"], r["document_url"], r["signal_type"],
             r["matched_term"], r["snippet"], r["source_id"],
             r["retrieved_on"]),
        )
        n += 1
    conn.commit()
    return n


def load_primary_buildout_signals(conn: sqlite3.Connection) -> int:
    p = DATA / "primary_buildout_signals.yaml"
    if not p.exists():
        return 0
    doc = read_yaml(p) or {}
    n = 0
    for r in doc.get("rows", []):
        conn.execute(
            """INSERT INTO primary_buildout_signals
               (signal_id, campus_id, reporting_company, company_bucket, ticker, counterparty,
                project_name, geography, claim_type, status_stage, capacity_basis,
                metric_value, metric_unit, as_of_date, expected_online_date,
                source_quote, notes, source_id)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["signal_id"], r.get("campus_id"), r["reporting_company"],
             r["company_bucket"], r.get("ticker"), r.get("counterparty"),
             r.get("project_name"), r.get("geography"), r["claim_type"],
             r["status_stage"], r.get("capacity_basis", "Unknown"),
             r.get("metric_value"), r.get("metric_unit"), r.get("as_of_date"),
             r.get("expected_online_date"), r.get("source_quote"),
             r.get("notes"), r["source_id"]),
        )
        n += 1
    conn.commit()
    return n


def calendar_quarter(as_of_date: object) -> str:
    """Calendar quarter 'YYYYQn' derived from a YYYY-MM-DD as_of_date.

    YAML may hand us a datetime.date (unquoted dates) or a string; both work.
    """
    if isinstance(as_of_date, (dt.date, dt.datetime)):
        as_of_date = as_of_date.isoformat()[:10]
    s = str(as_of_date)
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", s)
    if not m:
        raise ValueError(f"as_of_date requires YYYY-MM-DD, got {s!r}")
    month = int(m.group(2))
    if not 1 <= month <= 12:
        raise ValueError(f"invalid month in as_of_date {s!r}")
    return f"{m.group(1)}Q{((month - 1) // 3) + 1}"


def load_operator_disclosures(conn: sqlite3.Connection) -> tuple[int, int]:
    """Load operator capacity disclosures: terms registry + verbatim rows.

    terms: → operator_term_registry — roster, fiscal year ends, cross-table name
    map, the wording map, preferred bases, and the basis vocabulary. The wording
    map carries (stage_normalized, default_row_kind) ONLY; capacity_basis is
    row-level data and never part of the wording map.
    rows:  → operator_capacity_disclosures — as_of_quarter is derived here from
    as_of_date (calendar quarter), never hand-entered. source_id FK is enforced
    by SQLite, so a typo fails the load rather than passing silently.
    """
    p = DATA / "operator_disclosures.yaml"
    if not p.exists():
        return 0, 0
    doc = read_yaml(p) or {}
    terms = doc.get("terms") or {}
    n_reg = 0
    for token in terms.get("basis_vocabulary") or []:
        conn.execute(
            "INSERT INTO operator_term_registry (entry_kind, basis_token) VALUES ('basis', ?)",
            (token,),
        )
        n_reg += 1
    for op in terms.get("operators") or []:
        name = op["name"]
        conn.execute(
            """INSERT INTO operator_term_registry
               (entry_kind, operator, fiscal_year_end, pbs_name, campuses_name,
                disclosure_channel, cumulative_planned_flag)
               VALUES ('operator',?,?,?,?,?,?)""",
            (name, op.get("fiscal_year_end"), op.get("pbs_name"),
             op.get("campuses_name"), op.get("disclosure_channel"),
             int(bool(op.get("cumulative_planned", False)))),
        )
        n_reg += 1
        for t in op.get("terms") or []:
            conn.execute(
                """INSERT INTO operator_term_registry
                   (entry_kind, operator, stage_verbatim, stage_normalized, default_row_kind)
                   VALUES ('term',?,?,?,?)""",
                (name, t["verbatim"], t["stage"], t["row_kind"]),
            )
            n_reg += 1
        for pb in op.get("preferred_basis") or []:
            conn.execute(
                """INSERT INTO operator_term_registry
                   (entry_kind, operator, stage_normalized, preferred_basis)
                   VALUES ('preferred_basis',?,?,?)""",
                (name, pb["stage"], pb["basis"]),
            )
            n_reg += 1
    n_rows = 0
    for r in doc.get("rows") or []:
        as_of = r["as_of_date"]
        if isinstance(as_of, (dt.date, dt.datetime)):
            as_of = as_of.isoformat()[:10]
        as_of = str(as_of)
        conn.execute(
            """INSERT INTO operator_capacity_disclosures
               (disclosure_id, operator, operator_bucket, as_of_date, as_of_quarter,
                fiscal_label, stage_normalized, stage_verbatim, row_kind,
                component_label, mw_value, capacity_basis, tenant_operator,
                verbatim_quote, notes, source_id)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["disclosure_id"], r["operator"], r.get("operator_bucket"),
             as_of, calendar_quarter(as_of), r.get("fiscal_label"),
             r["stage_normalized"], r.get("stage_verbatim", ""),
             r["row_kind"], r.get("component_label", ""),
             r.get("mw_value"), r["capacity_basis"], r.get("tenant_operator"),
             r.get("verbatim_quote"), r.get("notes"), r.get("source_id")),
        )
        n_rows += 1
    conn.commit()
    return n_reg, n_rows


def load_qualitative_load_commentary(conn: sqlite3.Connection) -> int:
    p = DATA / "qualitative_load_commentary.yaml"
    if not p.exists():
        return 0
    doc = read_yaml(p) or {}
    n = 0
    for r in doc.get("rows", []):
        derived_bucket = timeline_bucket(r["statement_date"], r["date_precision"])
        supplied_bucket = r.get("timeline_bucket")
        if supplied_bucket is not None and supplied_bucket != derived_bucket:
            raise ValueError(
                f"qualitative_load_commentary {r['statement_id']} timeline_bucket "
                f"{supplied_bucket!r} != derived {derived_bucket!r}"
            )
        conn.execute(
            """INSERT INTO qualitative_load_commentary
               (statement_id, source_id, statement_date, date_precision, event_name,
                timeline_bucket, speaker_name, speaker_title, organization,
                organization_bucket, source_route, source_type, statement_taxonomy,
                polarity, load_stage, geography, related_company, short_quote,
                paraphrase, numeric_value, numeric_unit, capacity_basis,
                time_horizon_start, time_horizon_end, confidence,
                independence_group, notes)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["statement_id"], r["source_id"], r["statement_date"],
             r["date_precision"], r["event_name"], derived_bucket,
             r["speaker_name"], r.get("speaker_title"), r["organization"],
             r["organization_bucket"], r["source_route"], r["source_type"],
             r["statement_taxonomy"], r["polarity"], r["load_stage"],
             r.get("geography"), r.get("related_company"), r.get("short_quote"),
             r["paraphrase"], r.get("numeric_value"), r.get("numeric_unit"),
             r["capacity_basis"], r.get("time_horizon_start"),
             r.get("time_horizon_end"), r["confidence"],
             r["independence_group"], r.get("notes")),
        )
        n += 1
    conn.commit()
    return n


def load_lcoe(conn: sqlite3.Connection) -> int:
    doc = read_yaml(DATA / "lcoe" / "lcoe.yaml")
    n = 0
    for r in doc["rows"]:
        conn.execute(
            """INSERT INTO lcoe_data
               (technology, year_vintage, report_name, report_date, geography,
                subsidized, lcoe_low, lcoe_mid, lcoe_high, currency_year, notes, source_id)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r["technology"], r["year_vintage"], r["report_name"], r.get("report_date"),
             r.get("geography", "US"), int(r.get("subsidized", 0)),
             r.get("lcoe_low"), r.get("lcoe_mid"), r.get("lcoe_high"),
             r.get("currency_year"), r.get("notes"), r["source_id"]),
        )
        n += 1
    conn.commit()
    return n


def load_gas_capex(conn: sqlite3.Connection) -> int:
    doc = read_yaml(DATA / "capex" / "gas_capex.yaml")
    n = 0
    for r in doc["rows"]:
        conn.execute(
            """INSERT INTO gas_capex
               (year, plant_type, cost_low_kw, cost_mid_kw, cost_high_kw,
                data_type, label, notes, source_id)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (r["year"], r["plant_type"], r.get("cost_low_kw"), r.get("cost_mid_kw"),
             r.get("cost_high_kw"), r["data_type"], r["label"], r.get("notes"),
             r["source_id"]),
        )
        n += 1
    conn.commit()
    return n


def load_renewable_capex(conn: sqlite3.Connection) -> int:
    doc = read_yaml(DATA / "capex" / "renewable_capex.yaml")
    n = 0
    for r in doc["rows"]:
        conn.execute(
            """INSERT INTO renewable_capex
               (technology, year, cost_low_kw, cost_mid_kw, cost_high_kw,
                data_type, label, notes, source_id)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (r["technology"], r["year"], r.get("cost_low_kw"), r.get("cost_mid_kw"),
             r.get("cost_high_kw"), r["data_type"], r["label"], r.get("notes"),
             r["source_id"]),
        )
        n += 1
    conn.commit()
    return n


def load_turbine_supply(conn: sqlite3.Connection) -> int:
    doc = read_yaml(DATA / "capex" / "turbine_supply.yaml")
    n = 0
    for r in doc["rows"]:
        conn.execute(
            """INSERT INTO turbine_supply
               (manufacturer, as_of, backlog_note, notes, source_id)
               VALUES (?,?,?,?,?)""",
            (r["manufacturer"], r["as_of"], r["backlog_note"],
             r.get("notes"), r["source_id"]),
        )
        n += 1
    conn.commit()
    return n


def load_demand(conn: sqlite3.Connection) -> tuple[int, int, int]:
    doc = read_yaml(DATA / "demand" / "demand.yaml")
    n1 = n2 = n3 = 0
    for r in doc.get("demand_metrics", []):
        conn.execute(
            """INSERT INTO demand_metrics
               (metric, value_num, value_text, unit, year, geography, notes, source_id)
               VALUES (?,?,?,?,?,?,?,?)""",
            (r["metric"], r.get("value_num"), r.get("value_text"),
             r.get("unit"), r.get("year"), r.get("geography"),
             r.get("notes"), r["source_id"]),
        )
        n1 += 1
    for r in doc.get("grid_capacity_plan", []):
        conn.execute(
            """INSERT INTO grid_capacity_plan
               (technology, window_start, window_end, gw_planned, geography, notes, source_id)
               VALUES (?,?,?,?,?,?,?)""",
            (r["technology"], r["window_start"], r["window_end"], r["gw_planned"],
             r.get("geography", "US"), r.get("notes"), r["source_id"]),
        )
        n2 += 1
    for r in doc.get("hyperscaler_cumulative", []):
        conn.execute(
            """INSERT INTO hyperscaler_cumulative
               (company, as_of, metric, value_gw, notes, source_id)
               VALUES (?,?,?,?,?,?)""",
            (r["company"], r["as_of"], r["metric"], r["value_gw"],
             r.get("notes"), r["source_id"]),
        )
        n3 += 1
    conn.commit()
    return n1, n2, n3


def main() -> int:
    if DB.exists():
        DB.unlink()
    conn = sqlite3.connect(DB)
    try:
        rebuild_schema(conn)
        n_src = len(load_sources(conn))
        n_c = load_contracts(conn)
        n_cam = load_campuses(conn)
        n_ce = load_campus_evidence(conn)
        n_ps = load_proxy_signal_definitions(conn)
        n_spm = load_sec_proxy_metrics(conn)
        n_sfts = load_sec_filing_text_signals(conn)
        n_pbs = load_primary_buildout_signals(conn)
        n_otr, n_ocd = load_operator_disclosures(conn)
        n_pg = load_planned_generators(conn)
        n_og = load_operating_generators(conn)
        n_l = load_lcoe(conn)
        n_g = load_gas_capex(conn)
        n_r = load_renewable_capex(conn)
        n_t = load_turbine_supply(conn)
        n_d, n_p, n_h = load_demand(conn)
        n_qlc = load_qualitative_load_commentary(conn)
    except (sqlite3.IntegrityError, ValueError) as e:
        print(f"[LOAD FAILED] {type(e).__name__}: {e}", file=sys.stderr)
        print("Most common cause: a row references a source_id not present in sources.yaml.")
        return 1
    finally:
        conn.close()

    print(f"loaded: {n_src} sources")
    print(f"        {n_c} contracts")
    print(f"        {n_cam} campuses")
    print(f"        {n_ce} campus_evidence")
    print(f"        {n_ps} proxy_signal_definitions")
    print(f"        {n_spm} sec_proxy_metrics")
    print(f"        {n_sfts} sec_filing_text_signals")
    print(f"        {n_pbs} primary_buildout_signals")
    print(f"        {n_ocd} operator_capacity_disclosures")
    print(f"        {n_otr} operator_term_registry entries")
    print(f"        {n_qlc} qualitative_load_commentary")
    print(f"        {n_pg} planned_generators (EIA-860M)")
    print(f"        {n_og} operating_generators (EIA-860M latest)")
    print(f"        {n_l} lcoe")
    print(f"        {n_g} gas_capex")
    print(f"        {n_r} renewable_capex")
    print(f"        {n_t} turbine_supply")
    print(f"        {n_d} demand_metrics")
    print(f"        {n_p} grid_capacity_plan")
    print(f"        {n_h} hyperscaler_cumulative")
    print(f"db: {DB}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
